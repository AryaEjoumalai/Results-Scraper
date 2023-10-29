
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl.styles import PatternFill

driver = webdriver.Edge(r"msedgedriver.exe")



for roll in range(201,321):
    if True:
        #driver.get("http://exam.pondiuni.edu.in/results/result.php?r=21TD0"+"257"+"&e=D")
        driver.get("http://exam.pondiuni.edu.in/results/result.php?r=21TD0"+str(roll)+"&e=D")
        flag=False
        while True:
            
            try:
                driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[2]/td[7]/div")
                
            except:
                try:
                    not_found=driver.find_element(By.XPATH, "//*[@id='result_error_div']").text
                    flag=True
                    print(flag)
                    break
                except: continue
            break
        if flag==True: continue
        element=driver.find_element(By.XPATH, "//*[@id='student_info']/tbody/tr[3]/td")
        name=element.text[22:]
        math=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[2]/td[7]/div").text
        EDC=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[3]/td[7]/div").text
        OOPD=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[4]/td[7]/div").text
        DSD=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[5]/td[7]/div").text
        DS=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[6]/td[7]/div").text
        COA=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[7]/td[7]/div").text
        EDClab=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[8]/td[7]/div").text
        DSlab=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[9]/td[7]/div").text
        DSDlab=driver.find_element(By.XPATH, "//*[@id='results_subject_table']/tbody/tr[10]/td[7]/div").text
        
        workbook = openpyxl.load_workbook('results.xlsx')

        worksheet = workbook.active

        worksheet.cell(roll-199 , column=1).value = "21TD0"+str(roll)
        worksheet.cell(roll-199 , column=2).value = name
        
        worksheet.cell(roll-199 , column=3).value = math
        if math=='Fail': worksheet.cell(roll-199 , column=3).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        worksheet.cell(roll-199 , column=4).value = EDC
        if EDC=='Fail': worksheet.cell(roll-199 , column=4).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        worksheet.cell(roll-199 , column=5).value = OOPD
        if OOPD=='Fail': worksheet.cell(roll-199 , column=5).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        worksheet.cell(roll-199 , column=6).value = DSD
        if DSD=='Fail': worksheet.cell(roll-199 , column=6).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        worksheet.cell(roll-199 , column=7).value = DS
        if DS=='Fail': worksheet.cell(roll-199 , column=7).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        worksheet.cell(roll-199 , column=8).value = COA
        if COA=='Fail': worksheet.cell(roll-199 , column=8).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        worksheet.cell(roll-199 , column=9).value = EDClab
        if EDClab=='Fail': worksheet.cell(roll-199 , column=9).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        worksheet.cell(roll-199 , column=10).value = DSlab
        if DSlab=='Fail': worksheet.cell(roll-199 , column=10).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        worksheet.cell(roll-199 , column=11).value = DSDlab
        if DSDlab=='Fail': worksheet.cell(roll-199 , column=11).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        
        workbook.save('results.xlsx')
  

